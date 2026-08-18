import csv
import json
import logging
from collections import Counter, namedtuple
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

COR_CABECALHO = "173772"
COR_AMARELO = "FFF3CD"
COR_VERMELHO = "F8D7DA"
COR_BORDA_VERMELHA = "DC3545"

FONTE_CABECALHO = Font(bold=True, color="FFFFFF")
FUNDO_CABECALHO = PatternFill("solid", start_color=COR_CABECALHO)
FUNDO_AMARELO = PatternFill("solid", start_color=COR_AMARELO)
FUNDO_VERMELHO = PatternFill("solid", start_color=COR_VERMELHO)
BORDA_VERMELHA = Border(left=Side(style="medium", color=COR_BORDA_VERMELHA))

LARGURA_MINIMA = 12
LARGURA_MAXIMA = 30

"""
Forma intermediaria que o xlsx e o CSV consomem igual. Existir uma so
garante por construcao que os dois saem com as mesmas colunas, na mesma
ordem e com as mesmas linhas - em vez de ser coincidencia de dois codigos
parecidos que podem divergir depois. Os avisos so o xlsx usa, porque CSV
nao tem formatacao.
"""
Tabela = namedtuple("Tabela", "cabecalho linhas avisos")


# ---------------------------------------------------------------------------
# Pecas compartilhadas pelas duas planilhas
# ---------------------------------------------------------------------------


def escrever_cabecalho(aba, colunas):
    aba.append(colunas)

    for celula in aba[1]:
        celula.font = FONTE_CABECALHO
        celula.fill = FUNDO_CABECALHO


def escrever_linha(aba, numero_linha, valores, largura, aviso):
    """
    Escreve uma linha de dados ja com o destaque aplicado.

    Escreve celula a celula em vez de append() pra garantir que o
    preenchimento cubra a largura toda da linha, inclusive as colunas que
    ficam vazias - que na planilha do holerite sao a maioria.
    """
    # Vermelho ganha do amarelo quando os dois valem pra mesma linha.
    if aviso["nao_sequencial"]:
        fundo = FUNDO_VERMELHO
    elif aviso["amarelo"]:
        fundo = FUNDO_AMARELO
    else:
        fundo = None

    for numero_coluna in range(1, largura + 1):
        celula = aba.cell(row=numero_linha, column=numero_coluna)

        if numero_coluna <= len(valores):
            celula.value = valores[numero_coluna - 1]

        if fundo is not None:
            celula.fill = fundo

    if aviso["nao_sequencial"]:
        aba.cell(row=numero_linha, column=1).border = BORDA_VERMELHA


def ajustar_larguras(aba, cabecalho):
    for numero_coluna, titulo in enumerate(cabecalho, start=1):
        letra = aba.cell(row=1, column=numero_coluna).column_letter
        largura = max(LARGURA_MINIMA, min(len(titulo) + 2, LARGURA_MAXIMA))
        aba.column_dimensions[letra].width = largura


def tem_incerteza(textos):
    """O "?" do contrato marca caractere que nao deu pra ler com seguranca."""
    return any("?" in (texto or "") for texto in textos)


# ---------------------------------------------------------------------------
# Cartao de ponto
# ---------------------------------------------------------------------------


def linhas_do_cartao_ponto(dados):
    """
    Achata as paginas numa lista unica de dias, na ordem do documento.

    Nao ordena por data de proposito: a planilha tem que refletir a ordem em
    que os dias aparecem no PDF, senao uma data fora de lugar - que e
    justamente o que o aviso de nao-sequencial existe pra mostrar - sumiria
    depois de ordenar.
    """
    return [dia for pagina in dados["pages"] for dia in pagina["days"]]


def ler_data(date_raw):
    """
    "01/07/2012" -> date. Devolve None quando nao da pra ler, que e o caso de
    uma data com "?" de caractere ilegivel.
    """
    try:
        return datetime.strptime(date_raw, "%d/%m/%Y").date()
    except (ValueError, TypeError):
        return None


def contar_pares(linhas):
    """
    Quantos pares Entrada/Saida a planilha precisa, olhando o dia com mais
    batidas. Arredonda pra cima: um dia com 1 batida so ainda ocupa um par
    inteiro, com a Saida vazia.
    """
    maximo = max((len(linha["punches"]) for linha in linhas), default=0)
    return (maximo + 1) // 2


def montar_cabecalho_cartao_ponto(pares):
    colunas = ["Data"]

    for numero in range(1, pares + 1):
        colunas.append(f"Entrada {numero}")
        colunas.append(f"Saída {numero}")

    return colunas


def derivar_avisos_cartao_ponto(linhas):
    """
    Calcula os avisos de cada dia a partir do proprio dado, na hora de gerar.

    Nenhum deles e campo do JSON: batida impar sai de contar os punches, e
    nao-sequencial sai de comparar com a linha anterior.

    Uma data ilegivel nao quebra a cadeia: ela mesma nao vira aviso vermelho, e
    a proxima data legivel e comparada com a ultima legivel, nao com ela.
    """
    avisos = []
    ultima_data = None

    for linha in linhas:
        data = ler_data(linha["date_raw"])

        textos = [linha["date_raw"]]
        for punch in linha["punches"]:
            textos.append(punch["time_raw"])
            textos.append(punch["time_hhmm"])

        incerto = tem_incerteza(textos)
        impar = len(linha["punches"]) % 2 != 0

        if data is None or ultima_data is None:
            # Primeira linha, ou data ilegivel: nao ha com o que comparar.
            nao_sequencial = False
        else:
            nao_sequencial = (data - ultima_data).days != 1

        if data is not None:
            ultima_data = data

        avisos.append({
            "impar": impar,
            "incerto": incerto,
            "amarelo": impar or incerto,
            "nao_sequencial": nao_sequencial,
        })

    return avisos


def tabela_cartao_ponto(dados):
    """
    Monta a tabela do cartao de ponto: uma linha por dia, uma coluna Data mais
    os pares Entrada/Saida.
    """
    dias = linhas_do_cartao_ponto(dados)
    cabecalho = montar_cabecalho_cartao_ponto(contar_pares(dias))
    avisos = derivar_avisos_cartao_ponto(dias)

    linhas = [
        [dia["date_raw"]] + [punch["time_hhmm"] for punch in dia["punches"]]
        for dia in dias
    ]

    return Tabela(cabecalho, linhas, avisos)


def gerar_planilha_cartao_ponto(dados, caminho_saida):
    return escrever_xlsx(
        tabela_cartao_ponto(dados), "Cartão de ponto", caminho_saida, "A2"
    )


def gerar_csv_cartao_ponto(dados, caminho_saida):
    return escrever_csv(tabela_cartao_ponto(dados), caminho_saida)


# ---------------------------------------------------------------------------
# Holerite
# ---------------------------------------------------------------------------


def rotulos_das_verbas(pagina):
    """
    Nome de coluna de cada verba da pagina, na ordem em que aparecem.

    A mesma verba pode aparecer duas vezes na mesma folha com valores
    diferentes - "CONTRIBUICAO NEGOCIAL" lancada em duas parcelas, por
    exemplo. Como a planilha tem uma celula por coluna, a segunda ocorrencia
    ganha um contador no nome em vez de sobrescrever a primeira: o numero e
    marca nossa, mas perder o valor apagaria um dado que esta no documento.
    """
    vistos = Counter()
    rotulos = []

    for field in pagina["fields"]:
        vistos[field["label"]] += 1
        ocorrencia = vistos[field["label"]]

        if ocorrencia == 1:
            rotulos.append(field["label"])
        else:
            rotulos.append(f"{field['label']} ({ocorrencia})")

    return rotulos


def colunas_de_verbas(dados):
    """
    Uniao dos rotulos de verba, na ordem de primeira aparicao no documento.

    dict.fromkeys guarda a ordem de insercao e ignora repetido, que e
    exatamente "primeira aparicao vence".
    """
    return list(dict.fromkeys(
        rotulo
        for pagina in dados["pages"]
        for rotulo in rotulos_das_verbas(pagina)
    ))


def valores_por_verba(pagina):
    """Mapeia rotulo de coluna -> value dentro de uma pagina."""
    return {
        rotulo: field["value"]
        for rotulo, field in zip(rotulos_das_verbas(pagina), pagina["fields"])
    }


def tem_folhas(paginas):
    """
    Diz se o documento separa folhas dentro da pagina (mes, acerto).

    A coluna so aparece em quem tem folha rotulada: num holerite comum ela
    ficaria vazia da primeira a ultima linha.
    """
    return any(pagina.get("folha") for pagina in paginas)


def ler_competencia(pagina):
    """
    Vira a competencia num numero absoluto de meses (ano * 12 + mes).

    Compara em meses absolutos pra dezembro -> janeiro contar como
    consecutivo. Devolve None quando a competencia nao deu pra ler.
    """
    try:
        return int(pagina["year"]) * 12 + int(pagina["month"])
    except (ValueError, TypeError):
        return None


def derivar_avisos_holerite(paginas):
    """
    Calcula os avisos de cada pagina a partir do proprio dado.

    Pagina vazia e a que nao rendeu verba nenhuma - a linha dela sai so com
    Pag./Mes/Ano e o resto em branco, entao precisa saltar aos olhos.

    Competencia ilegivel nao quebra a cadeia: ela mesma nao vira vermelho, e a
    proxima legivel e comparada com a ultima legivel, nao com ela.
    """
    avisos = []
    ultima_competencia = None

    for pagina in paginas:
        competencia = ler_competencia(pagina)

        textos = [pagina["month"], pagina["year"]]
        textos += [field["value"] for field in pagina["fields"]]

        vazia = not pagina["fields"]
        incerto = tem_incerteza(textos)

        if competencia is None or ultima_competencia is None:
            nao_sequencial = False
        else:
            """
            Repetir a competencia nao e furo: duas folhas da mesma pagina (mes
            e acerto) sao duas linhas do mesmo mes de proposito. So pular ou
            voltar no tempo e que merece o vermelho.
            """
            nao_sequencial = competencia - ultima_competencia not in (0, 1)

        if competencia is not None:
            ultima_competencia = competencia

        avisos.append({
            "vazia": vazia,
            "incerto": incerto,
            "amarelo": vazia or incerto,
            "nao_sequencial": nao_sequencial,
        })

    return avisos


def tabela_holerite(dados):
    """
    Monta a tabela do holerite: uma linha por pagina, colunas fixas Pag./Mes/
    Ano e depois uma coluna por verba distinta.

    O documento e uma lista vertical de verbas por pagina; a tabela e uma
    matriz larga. Essa transposicao e o trabalho.
    """
    paginas = dados["pages"]
    verbas = colunas_de_verbas(dados)
    com_folha = tem_folhas(paginas)

    fixas = ["Pág."] + (["Folha"] if com_folha else []) + ["Mês", "Ano"]
    cabecalho = fixas + verbas
    avisos = derivar_avisos_holerite(paginas)

    linhas = []

    for pagina in paginas:
        valores_da_pagina = valores_por_verba(pagina)

        """
        Mes e Ano ficam string, como vieram do contrato: virar numero comeria
        o zero a esquerda de "01" e a saida mostraria 1.
        """
        linha = [pagina["page"]]

        if com_folha:
            # Duas folhas da mesma pagina viram duas linhas com o mesmo
            # numero: e a folha que diz qual e qual.
            linha.append(pagina.get("folha", ""))

        linha += [pagina["month"], pagina["year"]]

        # None deixa a celula em branco de verdade quando a verba nao aparece
        # nesta pagina; "" criaria uma string vazia.
        linha += [valores_da_pagina.get(verba) for verba in verbas]

        linhas.append(linha)

    return Tabela(cabecalho, linhas, avisos)


def gerar_planilha_holerite(dados, caminho_saida):
    """
    O congelamento acompanha as colunas fixas: sem folha sao Pag./Mes/Ano e a
    rolagem comeca em D; com folha entra mais uma e comeca em E.
    """
    congelar = "E2" if tem_folhas(dados["pages"]) else "D2"

    return escrever_xlsx(
        tabela_holerite(dados), "Holerite", caminho_saida, congelar
    )


def gerar_csv_holerite(dados, caminho_saida):
    return escrever_csv(tabela_holerite(dados), caminho_saida)


# ---------------------------------------------------------------------------
# Escrita dos formatos
# ---------------------------------------------------------------------------


def escrever_xlsx(tabela, titulo_aba, caminho_saida, congelar):
    planilha = Workbook()
    aba = planilha.active
    aba.title = titulo_aba

    escrever_cabecalho(aba, tabela.cabecalho)

    for numero_linha, (valores, aviso) in enumerate(
        zip(tabela.linhas, tabela.avisos), start=2
    ):
        escrever_linha(aba, numero_linha, valores, len(tabela.cabecalho), aviso)

    aba.freeze_panes = congelar
    ajustar_larguras(aba, tabela.cabecalho)
    planilha.save(caminho_saida)

    logger.debug(
        "xlsx %s: %s linhas, %s colunas, %s destaques",
        caminho_saida,
        len(tabela.linhas),
        len(tabela.cabecalho),
        sum(1 for a in tabela.avisos if a["amarelo"] or a["nao_sequencial"]),
    )

    return caminho_saida


def escrever_csv(tabela, caminho_saida):
    """
    Escreve os mesmos dados do xlsx - mesmas colunas, mesma ordem, mesmas
    linhas -, sem os destaques, porque CSV nao tem formatacao.

    Separador ";" e nao ",": o decimal brasileiro ja usa virgula, entao com
    "," todo valor sairia entre aspas. E o Excel em portugues espera ";" como
    separador de lista, senao joga a linha inteira na primeira coluna.

    utf-8-sig grava o BOM na frente do arquivo: sem ele o Excel abre o CSV no
    encoding do sistema e a acentuacao sai trocada.

    newline vazio e exigencia do modulo csv, que ja escreve a quebra de linha
    sozinho; sem isso o Windows dobra a quebra e sai uma linha em branco entre
    cada duas.
    """
    with open(caminho_saida, "w", encoding="utf-8-sig", newline="") as arquivo:
        escritor = csv.writer(arquivo, delimiter=";")
        escritor.writerow(tabela.cabecalho)

        for valores in tabela.linhas:
            # Completa a largura igual o xlsx faz: linha curta vira celula
            # vazia no fim, nao coluna faltando.
            faltam = len(tabela.cabecalho) - len(valores)
            escritor.writerow(list(valores) + [None] * faltam)

    logger.debug(
        "csv %s: %s linhas, %s colunas",
        caminho_saida,
        len(tabela.linhas),
        len(tabela.cabecalho),
    )

    return caminho_saida


def gerar_json(dados, caminho_saida):
    """
    Grava o proprio value da transcricao, sem transformacao nenhuma.

    Serve pros dois tipos de documento: o que sai e exatamente o que entrou.
    ensure_ascii=False mantem os acentos legiveis em vez de virar escape
    unicode.
    """
    with open(caminho_saida, "w", encoding="utf-8") as arquivo:
        json.dump(dados, arquivo, indent=2, ensure_ascii=False)

    return caminho_saida


"""
Escolha do gerador por (tipo, formato). As tres funcoes de cada tipo tem a
mesma assinatura (dados, caminho_saida), entao a rota de download so procura
aqui e chama, sem um if por formato.
"""
GERADORES = {
    "cartao-ponto": {
        "xlsx": gerar_planilha_cartao_ponto,
        "csv": gerar_csv_cartao_ponto,
        "json": gerar_json,
    },
    "holerite": {
        "xlsx": gerar_planilha_holerite,
        "csv": gerar_csv_holerite,
        "json": gerar_json,
    },
}
